import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('用车费用明细.xlsx')
sheet = wb.active

print("Verifying Photo Log Row 27 to 48:")
print(f"Row 27: {[sheet.cell(row=27, column=c).value for c in range(2, 7)]}")
for r in range(28, 49):
    row_vals = [sheet.cell(row=r, column=c).value for c in range(2, 7)]
    # Also count images in this row if any
    row_images = [img for img in sheet._images if img.anchor == f"B{r}"]
    has_img = len(row_images) > 0
    print(f"Row {r}: {row_vals} | Has Image: {has_img}")
