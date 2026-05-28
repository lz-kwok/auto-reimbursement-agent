import openpyxl
import datetime
import sys

sys.stdout.reconfigure(encoding='utf-8')

# The 16 legs data
legs = [
    {"date": datetime.date(2026, 4, 2), "start_city": "南京市", "start_odo": 170760, "end_city": "合肥市", "end_odo": 170958},
    {"date": datetime.date(2026, 4, 3), "start_city": "合肥市", "start_odo": 170958, "end_city": "南京市", "end_odo": 171167},
    {"date": datetime.date(2026, 4, 6), "start_city": "南京市", "start_odo": 171257, "end_city": "马鞍山市", "end_odo": 171328},
    {"date": datetime.date(2026, 4, 7), "start_city": "马鞍山市", "start_odo": 171328, "end_city": "南京市", "end_odo": 171397},
    {"date": datetime.date(2026, 4, 7), "start_city": "南京市", "start_odo": 171397, "end_city": "马鞍山市", "end_odo": 171468},
    {"date": datetime.date(2026, 4, 8), "start_city": "马鞍山市", "start_odo": 171468, "end_city": "南京市", "end_odo": 171538},
    {"date": datetime.date(2026, 4, 8), "start_city": "南京市", "start_odo": 171538, "end_city": "马鞍山市", "end_odo": 171608},
    {"date": datetime.date(2026, 4, 9), "start_city": "马鞍山市", "start_odo": 171608, "end_city": "南京市", "end_odo": 171701},
    {"date": datetime.date(2026, 4, 17), "start_city": "南京市", "start_odo": 172237, "end_city": "合肥市", "end_odo": 172431},
    {"date": datetime.date(2026, 4, 17), "start_city": "合肥市", "start_odo": 172431, "end_city": "南京市", "end_odo": 172626},
    {"date": datetime.date(2026, 4, 22), "start_city": "南京市", "start_odo": 172973, "end_city": "合肥市", "end_odo": 173161},
    {"date": datetime.date(2026, 4, 23), "start_city": "合肥市", "start_odo": 173161, "end_city": "南京市", "end_odo": 173357},
    {"date": datetime.date(2026, 4, 23), "start_city": "南京市", "start_odo": 173357, "end_city": "合肥市", "end_odo": 173546},
    {"date": datetime.date(2026, 4, 24), "start_city": "合肥市", "start_odo": 173546, "end_city": "南京市", "end_odo": 173738},
    {"date": datetime.date(2026, 4, 27), "start_city": "南京市", "start_odo": 173900, "end_city": "合肥市", "end_odo": 174053},
    {"date": datetime.date(2026, 4, 28), "start_city": "合肥市", "start_odo": 174053, "end_city": "南京市", "end_odo": 174258}
]

# Step 1: Remove merged cells and save to clear the internal cache
wb = openpyxl.load_workbook('用车费用明细.xlsx')
sheet = wb.active

ranges_to_remove = []
for r in list(sheet.merged_cells.ranges):
    if r.min_row >= 13:
        ranges_to_remove.append(r)

for r in ranges_to_remove:
    print(f"Removing incorrect merged range: {r}")
    sheet.merged_cells.remove(r)

wb.save('用车费用明细.xlsx')

# Step 2: Reload workbook and populate values
wb = openpyxl.load_workbook('用车费用明细.xlsx')
sheet = wb.active

# Fill data rows 5 to 20
for idx, leg in enumerate(legs):
    r = 5 + idx
    sheet.cell(row=r, column=1, value=leg['date'])
    sheet.cell(row=r, column=3, value=leg['start_city'])
    sheet.cell(row=r, column=4, value=leg['start_odo'])
    sheet.cell(row=r, column=5, value=leg['end_city'])
    sheet.cell(row=r, column=6, value=leg['end_odo'])
    sheet.cell(row=r, column=8, value=f"=F{r}-D{r}")
    sheet.cell(row=r, column=9, value=f"=H{r}*0.8")

# Step 3: Re-add correct merged cells
print("Adding merged range C22:D22")
sheet.merge_cells("C22:D22")

print("Adding merged range G25:I25")
sheet.merge_cells("G25:I25")

# Update formulas in row 21 (合计：)
sheet.cell(row=21, column=8, value="=SUM(H5:H20)")
sheet.cell(row=21, column=9, value="=SUM(I5:I20)")
sheet.cell(row=21, column=11, value="=SUM(K5:K20)")

# Update formulas in row 25 (summary details)
sheet.cell(row=25, column=2, value="=H21") # 总里程数
sheet.cell(row=25, column=3, value="=H21") # 实报里程数
sheet.cell(row=25, column=5, value="=K21") # 过路费
sheet.cell(row=25, column=7, value="=C25*0.8+E25+F25") # 报销金额

wb.save('用车费用明细.xlsx')
print("Successfully fixed Excel cells and merged cells.")
