import openpyxl
import datetime
import sys
import os
import re
from PIL import Image
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout.reconfigure(encoding='utf-8')

def parse_gps(ocr_raw):
    lat = None
    lng = None
    
    # 1. Find latitude
    for idx, text in enumerate(ocr_raw):
        if any(w in text for w in ['北纬', 'J纬', '纬']):
            for offset in [1, 2]:
                if idx + offset < len(ocr_raw):
                    val = ocr_raw[idx + offset]
                    digits = "".join(re.findall(r'\d+', val))
                    if len(digits) in [4, 5]:
                        if digits.startswith('8'):
                            digits = '3' + digits[1:]
                        if len(digits) == 5:
                            lat = f"北纬 {digits[:2]}°{digits[3:]}'"
                        elif len(digits) == 4:
                            lat = f"北纬 {digits[:2]}°{digits[2:]}'"
                        break
            if lat:
                break
                
    # 2. Find longitude
    for idx, text in enumerate(ocr_raw):
        if any(w in text for w in ['东经', '经']):
            for offset in [1, 2]:
                if idx + offset < len(ocr_raw):
                    val = ocr_raw[idx + offset]
                    digits = "".join(re.findall(r'\d+', val))
                    if len(digits) in [5, 6]:
                        if len(digits) == 6:
                            lng = f"东经 {digits[:3]}°{digits[4:]}'"
                        elif len(digits) == 5:
                            lng = f"东经 {digits[:3]}°{digits[3:]}'"
                        break
            if lng:
                break
                
    if lat and lng:
        return f"{lat}, {lng}"
    elif lat:
        return lat
    elif lng:
        return lng
    return ""

def get_clean_odometer(ocr_raw):
    for text in ocr_raw:
        match = re.search(r'17\d{4,5}', text)
        if match:
            val_str = match.group(0)
            if len(val_str) >= 6:
                return int(val_str[:6])
            elif len(val_str) == 5:
                return int(val_str) * 10
    for text in ocr_raw:
        digits = re.findall(r'\d+', text)
        for d in digits:
            if d.startswith('17'):
                if len(d) >= 6:
                    return int(d[:6])
                else:
                    return int(d)
    return None

# 1. Load extracted OCR data
with open('scratch/extracted_data.json', 'r', encoding='utf-8') as f:
    images_data = json = __import__('json').load(f)

# Update odometer, parse date, and parse GPS
gps_by_location = {}
for item in images_data:
    item['odo'] = get_clean_odometer(item['ocr_raw'])
    if item['date']:
        parts = item['date'].split('.')
        item['date_obj'] = datetime.date(int(parts[0]), int(parts[1]), int(parts[2]))
    else:
        item['date_obj'] = None
    
    # Parse GPS raw
    item['gps_raw'] = parse_gps(item['ocr_raw'])
    loc = item['location']
    if loc and item['gps_raw'] and ',' in item['gps_raw']:
        gps_by_location[loc] = item['gps_raw']

# Reconcile missing GPS components by matching locations
for item in images_data:
    loc = item['location']
    if loc and loc in gps_by_location:
        item['gps'] = gps_by_location[loc]
    else:
        item['gps'] = item['gps_raw']

# Sort images by odometer (chronological/mileage sequence)
images_data.sort(key=lambda x: x['odo'] if x['odo'] else 0)

wb = openpyxl.load_workbook('用车费用明细.xlsx')
sheet = wb.active

# 2. Setup styles
font_bold = Font(name="宋体", size=11, bold=True)
font_regular = Font(name="宋体", size=11)
fill_header = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")
align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")

thin_side = Side(border_style="thin", color="D9D9D9")
border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# 3. Setup Photo Log Header at Row 27
start_row = 27
sheet.row_dimensions[start_row].height = 25

# Clean old date header from F27 and any others in G27
sheet.cell(row=start_row, column=6).value = None
sheet.cell(row=start_row, column=7).value = None

headers = {
    2: "车辆仪表照片",
    3: "对应图片文件名",
    4: "仪表盘总里程 (km)",
    5: "照片水印地点",
    6: "经纬度",
    7: "拍摄日期"
}

for col, text in headers.items():
    cell = sheet.cell(row=start_row, column=col, value=text)
    cell.font = font_bold
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_all

# Set column widths
sheet.column_dimensions['B'].width = 22
sheet.column_dimensions['C'].width = 20
sheet.column_dimensions['D'].width = 20
sheet.column_dimensions['E'].width = 30
sheet.column_dimensions['F'].width = 30
sheet.column_dimensions['G'].width = 15

# Clean any existing date values from column F and G in the photos log region
for r in range(start_row + 1, start_row + 50):
    sheet.cell(row=r, column=6).value = None
    sheet.cell(row=r, column=7).value = None

# 4. Process and insert images starting from Row 28
for idx, item in enumerate(images_data):
    r = start_row + 1 + idx
    sheet.row_dimensions[r].height = 165  # 165pt corresponds to ~220px height
    
    # Path setup
    img_path = item['file']
    comp_path = os.path.join('compressed', img_path)
    
    # Compress image to 15% width and height (if not already done)
    if not os.path.exists(comp_path):
        with Image.open(img_path) as pil_img:
            width, height = pil_img.size
            new_width = int(width * 0.15)
            new_height = int(height * 0.15)
            resized_img = pil_img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            resized_img.save(comp_path, quality=70)
        
    # Re-insert image (overwrites old image placement if any, or openpyxl appends it)
    excel_img = OpenpyxlImage(comp_path)
    sheet.add_image(excel_img, f"B{r}")
    
    # Write metadata
    c_file = sheet.cell(row=r, column=3, value=item['file'])
    c_odo = sheet.cell(row=r, column=4, value=item['odo'])
    c_loc = sheet.cell(row=r, column=5, value=item['location'])
    c_gps = sheet.cell(row=r, column=6, value=item['gps'])
    c_date = sheet.cell(row=r, column=7, value=item['date_obj'])
    
    # Styling
    for col in [2, 3, 4, 5, 6, 7]:
        cell = sheet.cell(row=r, column=col)
        cell.font = font_regular
        cell.border = border_all
        if col in [5, 6]:
            cell.alignment = align_left
        else:
            cell.alignment = align_center
            
    c_date.number_format = 'yyyy-mm-dd'
    c_odo.number_format = '#,##0'

wb.save('用车费用明细.xlsx')
print(f"Successfully processed GPS coordinates and updated 用车费用明细.xlsx.")
