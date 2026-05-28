import openpyxl
import datetime
import sys

sys.stdout.reconfigure(encoding='utf-8')

# The 16 legs data
legs = [
    {"date": datetime.date(2026, 4, 2), "start_city": "南京市", "start_odo": 170760, "end_city": "合肥市", "end_odo": 170958},
    {"date": datetime.date(2026, 4, 3), "start_city": "合肥市", "start_odo": 170958, "end_city": "南京市", "end_odo": 171167},
    {"date": datetime.date(2026, 4, 6), "start_city": "南京市", "start_odo": 171257, "end_city": "马鞍山市", "end_odo": 171328},
    {"date": datetime.date(2026, 4, 7), "start_city": "马鞍山市", "start_odo": 171328, "end_city": "南京市", "end_odo": 171397},
    {"date": datetime.date(2026, 4, 7), "start_city": "南京市", "start_odo": 171397, "end_city": "马鞍山市", "end_odo": 171468},
    {"date": datetime.date(2026, 4, 8), "start_city": "马鞍山市", "start_odo": 171468, "end_city": "南京市", "end_odo": 171538},
    {"date": datetime.date(2026, 4, 8), "start_city": "南京市", "start_odo": 171538, "end_city": "马鞍山市", "end_odo": 171608},
    {"date": datetime.date(2026, 4, 9), "start_city": "马鞍山市", "start_odo": 171608, "end_city": "南京市", "end_odo": 171701},
    {"date": datetime.date(2026, 4, 17), "start_city": "南京市", "start_odo": 172237, "end_city": "合肥市", "end_odo": 172431},
    {"date": datetime.date(2026, 4, 17), "start_city": "合肥市", "start_odo": 172431, "end_city": "南京市", "end_odo": 172626},
    {"date": datetime.date(2026, 4, 22), "start_city": "南京市", "start_odo": 172973, "end_city": "合肥市", "end_odo": 173161},
    {"date": datetime.date(2026, 4, 23), "start_city": "合肥市", "start_odo": 173161, "end_city": "南京市", "end_odo": 173357},
    {"date": datetime.date(2026, 4, 23), "start_city": "南京市", "start_odo": 173357, "end_city": "合肥市", "end_odo": 173546},
    {"date": datetime.date(2026, 4, 24), "start_city": "合肥市", "start_odo": 173546, "end_city": "南京市", "end_odo": 173738},
    {"date": datetime.date(2026, 4, 27), "start_city": "南京市", "start_odo": 173900, "end_city": "合肥市", "end_odo": 174053},
    {"date": datetime.date(2026, 4, 28), "start_city": "合肥市", "start_odo": 174053, "end_city": "南京市", "end_odo": 174258}
]

wb = openpyxl.load_workbook('用车费用明细.xlsx')
sheet = wb.active

# 1. Fill data rows from 5 to 20
for idx, leg in enumerate(legs):
    r = 5 + idx
    sheet.cell(row=r, column=1, value=leg['date'])
    sheet.cell(row=r, column=3, value=leg['start_city'])
    sheet.cell(row=r, column=4, value=leg['start_odo'])
    sheet.cell(row=r, column=5, value=leg['end_city'])
    sheet.cell(row=r, column=6, value=leg['end_odo'])
    # Formulas
    sheet.cell(row=r, column=8, value=f"=F{r}-D{r}")
    sheet.cell(row=r, column=9, value=f"=H{r}*0.8")

# 2. Find and update the total row (合计：)
total_row_idx = None
for r in range(5, 50):
    if sheet.cell(row=r, column=2).value == "合计：":
        total_row_idx = r
        break

print(f"Updating total row at Row {total_row_idx}")
sheet.cell(row=total_row_idx, column=8, value=f"=SUM(H5:H{total_row_idx-1})")
sheet.cell(row=total_row_idx, column=9, value=f"=SUM(I5:I{total_row_idx-1})")
sheet.cell(row=total_row_idx, column=11, value=f"=SUM(K5:K{total_row_idx-1})")

# 3. Find and update summary row (below total row)
summary_header_row = None
for r in range(total_row_idx, total_row_idx + 10):
    if sheet.cell(row=r, column=2).value == "总里程数":
        summary_header_row = r
        break

print(f"Updating summary data row at Row {summary_header_row + 1}")
summary_data_row = summary_header_row + 1
sheet.cell(row=summary_data_row, column=2, value=f"=H{total_row_idx}") # 总里程数
sheet.cell(row=summary_data_row, column=3, value=f"=H{total_row_idx}") # 实报里程数
sheet.cell(row=summary_data_row, column=5, value=f"=K{total_row_idx}") # 过路费
sheet.cell(row=summary_data_row, column=7, value=f"=C{summary_data_row}*0.8+E{summary_data_row}+F{summary_data_row}") # 报销金额

wb.save('用车费用明细.xlsx')
print("Successfully corrected Excel file.")
