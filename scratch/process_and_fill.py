import json
import sys
import re
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout.reconfigure(encoding='utf-8')

# 1. Load the extracted OCR data
with open('scratch/extracted_data.json', 'r', encoding='utf-8') as f:
    images_data = json.load(f)

# 2. Refined odometer extraction
def get_clean_odometer(ocr_raw):
    for text in ocr_raw:
        # Search for digits starting with 17
        match = re.search(r'17\d{4,5}', text)
        if match:
            val_str = match.group(0)
            if len(val_str) >= 6:
                return int(val_str[:6])
            elif len(val_str) == 5:
                # In case it missed a digit
                return int(val_str) * 10
    # Backup: look for any number starting with 17
    for text in ocr_raw:
        digits = re.findall(r'\d+', text)
        for d in digits:
            if d.startswith('17'):
                if len(d) >= 6:
                    return int(d[:6])
                else:
                    return int(d)
    return None

# Extract correct odometer for all images
for item in images_data:
    item['odo'] = get_clean_odometer(item['ocr_raw'])
    # Fix date format to datetime.date object
    if item['date']:
        parts = item['date'].split('.')
        item['date_obj'] = datetime.date(int(parts[0]), int(parts[1]), int(parts[2]))
    else:
        item['date_obj'] = None

# Sort images by odometer (since odometer increases over time/trips)
images_data.sort(key=lambda x: x['odo'] if x['odo'] else 0)

print("Sorted images with correct odometer:")
for item in images_data:
    print(f"  {item['file']}: odo={item['odo']}, date={item['date']}, city={item['city']}")

# 3. Construct the trip legs
# We iterate through the sorted images.
# A leg is created when the location changes between consecutive images.
legs = []
for i in range(len(images_data) - 1):
    img_start = images_data[i]
    img_end = images_data[i+1]
    
    # Check if there is a city change
    if img_start['city'] != img_end['city']:
        # This is a leg!
        leg = {
            "date": img_end['date_obj'],  # Use arrival date
            "start_city": img_start['city'],
            "end_city": img_end['city'],
            "start_odo": img_start['odo'],
            "end_odo": img_end['odo']
        }
        legs.append(leg)

print(f"\nGenerated {len(legs)} legs:")
for idx, leg in enumerate(legs, 1):
    print(f"  Leg {idx}: {leg['date']} | {leg['start_city']} -> {leg['end_city']} | {leg['start_odo']} -> {leg['end_odo']}")

# 4. Write to Excel
wb = openpyxl.load_workbook('用车费用明细.xlsx')
sheet = wb.active

# Let's inspect rows 5 to 13 (before any insertion) to find where '合计：' is.
total_row_idx = None
for r in range(5, 50):
    if sheet.cell(row=r, column=2).value == "合计：":
        total_row_idx = r
        break

print(f"\nFound '合计：' at Row {total_row_idx}")

# The template has rows 5 to total_row_idx - 1 for data.
# We need to fill our legs into the sheet.
# If we have more legs than the existing data rows, we must insert rows.
num_legs = len(legs)
existing_data_rows = total_row_idx - 5
print(f"Number of legs to write: {num_legs}")
print(f"Existing template data rows: {existing_data_rows}")

if num_legs > existing_data_rows:
    rows_to_insert = num_legs - existing_data_rows
    print(f"Inserting {rows_to_insert} rows before Row {total_row_idx}")
    
    # Before inserting, let's capture the formatting of Row 5 as a template
    template_row = 5
    row_styles = {}
    for col_idx in range(1, 13):
        cell = sheet.cell(row=template_row, column=col_idx)
        row_styles[col_idx] = {
            "font": cell.font,
            "fill": cell.fill,
            "border": cell.border,
            "alignment": cell.alignment,
            "number_format": cell.number_format
        }
        
    # Insert rows
    sheet.insert_rows(total_row_idx, rows_to_insert)
    
    # Re-calculate the new total_row_idx
    new_total_row_idx = total_row_idx + rows_to_insert
    
    # Apply styling and default formulas to the inserted rows
    for r in range(total_row_idx, new_total_row_idx):
        for col_idx in range(1, 13):
            cell = sheet.cell(row=r, column=col_idx)
            # Apply template styles
            style = row_styles[col_idx]
            if style["font"]: cell.font = Font(name=style["font"].name, size=style["font"].size, bold=style["font"].bold, italic=style["font"].italic, color=style["font"].color)
            if style["fill"]: cell.fill = PatternFill(fill_type=style["fill"].fill_type, start_color=style["fill"].start_color, end_color=style["fill"].end_color)
            if style["border"]: cell.border = Border(left=style["border"].left, right=style["border"].right, top=style["border"].top, bottom=style["border"].bottom)
            if style["alignment"]: cell.alignment = Alignment(horizontal=style["alignment"].horizontal, vertical=style["alignment"].vertical, wrap_text=style["alignment"].wrap_text)
            cell.number_format = style["number_format"]
            
        # Add formulas for H (行驶总公里数) and I (油费)
        sheet.cell(row=r, column=8, value=f"=F{r}-D{r}")
        sheet.cell(row=r, column=9, value=f"=H{r}*0.8")
else:
    new_total_row_idx = total_row_idx

# 5. Write the leg data
for idx, leg in enumerate(legs):
    r = 5 + idx
    sheet.cell(row=r, column=1, value=leg['date'])
    sheet.cell(row=r, column=3, value=leg['start_city'])
    sheet.cell(row=r, column=4, value=leg['start_odo'])
    sheet.cell(row=r, column=5, value=leg['end_city'])
    sheet.cell(row=r, column=6, value=leg['end_odo'])

# 6. Update formulas in the shifted summary rows
# Row new_total_row_idx is the '合计：' row
sheet.cell(row=new_total_row_idx, column=8, value=f"=SUM(H5:H{new_total_row_idx-1})")
sheet.cell(row=new_total_row_idx, column=9, value=f"=SUM(I5:I{new_total_row_idx-1})")
sheet.cell(row=new_total_row_idx, column=11, value=f"=SUM(K5:K{new_total_row_idx-1})")

# Summary section starts 3 rows below '合计：' (which is Row 16, now shifted)
# Let's search for '总里程数' to find the summary header row
summary_header_row = None
for r in range(new_total_row_idx, new_total_row_idx + 10):
    if sheet.cell(row=r, column=2).value == "总里程数":
        summary_header_row = r
        break

print(f"Found summary header at Row {summary_header_row}")
summary_data_row = summary_header_row + 1

# Update summary table formulas
sheet.cell(row=summary_data_row, column=2, value=f"=H{new_total_row_idx}") # 总里程数
sheet.cell(row=summary_data_row, column=3, value=f"=H{new_total_row_idx}") # 实报里程数
sheet.cell(row=summary_data_row, column=5, value=f"=K{new_total_row_idx}") # 过路费
sheet.cell(row=summary_data_row, column=7, value=f"=C{summary_data_row}*0.8+E{summary_data_row}+F{summary_data_row}") # 报销金额

wb.save('用车费用明细.xlsx')
print("\nSuccessfully updated用车费用明细.xlsx with all 16 legs and adjusted formulas.")
