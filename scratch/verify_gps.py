import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook('用车费用明细.xlsx')
sheet = wb.active

print("Verifying Photo Log Row 27 to 32:")
print(f"Row 27: {[sheet.cell(row=27, column=c).value for c in range(2, 8)]}")
for r in range(28, 33):
    row_vals = [sheet.cell(row=r, column=c).value for c in range(2, 8)]
    print(f"Row {r}: {row_vals}")
