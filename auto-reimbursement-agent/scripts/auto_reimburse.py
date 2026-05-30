import os
import sys
import re
import datetime
import json
import urllib.parse
from PIL import Image
from PIL.ExifTags import TAGS
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def get_location_from_exif(img_path):
    try:
        img = Image.open(img_path)
        exif = img._getexif()
        if exif:
            for tag, val in exif.items():
                tag_name = TAGS.get(tag, tag)
                if tag_name == 'UserComment':
                    comment_str = ""
                    if isinstance(val, bytes):
                        for encoding in ['utf-8', 'gbk', 'latin-1']:
                            try:
                                comment_str = val.decode(encoding)
                                break
                            except Exception:
                                pass
                    else:
                        comment_str = str(val)
                    
                    if "StoryCamera=" in comment_str:
                        json_part = comment_str.split("StoryCamera=")[1].strip()
                        decoded_json = urllib.parse.unquote(json_part)
                        decoded_json = re.sub(r'[\x00-\x1f]', '', decoded_json)
                        try:
                            data = json.loads(decoded_json)
                            pos_name = data.get("pos", {}).get("name", "")
                            if pos_name:
                                return pos_name
                        except Exception:
                            pass
    except Exception:
        pass
    return None

def get_clean_odometer(ocr_raw):
    for text in ocr_raw:
        match = re.search(r'17\d{4,5}', text)
        if match:
            val_str = match.group(0)
            if len(val_str) >= 6:
                return int(val_str[:6])
            elif len(val_str) == 5:
                return int(val_str) * 10
    for text in ocr_raw:
        digits = re.findall(r'\d+', text)
        for d in digits:
            if d.startswith('17'):
                if len(d) >= 6:
                    return int(d[:6])
                else:
                    return int(d)
    return None

def parse_gps(ocr_raw):
    lat = None
    lng = None
    
    # 1. Find latitude
    for idx, text in enumerate(ocr_raw):
        if any(w in text for w in ['北纬', 'J纬', '纬']):
            for offset in [1, 2]:
                if idx + offset < len(ocr_raw):
                    val = ocr_raw[idx + offset]
                    digits = "".join(re.findall(r'\d+', val))
                    if len(digits) in [4, 5]:
                        if digits.startswith('8'):
                            digits = '3' + digits[1:]
                        if len(digits) == 5:
                            lat = f"北纬 {digits[:2]}°{digits[3:]}'"
                        elif len(digits) == 4:
                            lat = f"北纬 {digits[:2]}°{digits[2:]}'"
                        break
            if lat:
                break
                
    # 2. Find longitude
    for idx, text in enumerate(ocr_raw):
        if any(w in text for w in ['东经', '经']):
            for offset in [1, 2]:
                if idx + offset < len(ocr_raw):
                    val = ocr_raw[idx + offset]
                    digits = "".join(re.findall(r'\d+', val))
                    if len(digits) in [5, 6]:
                        if len(digits) == 6:
                            lng = f"东经 {digits[:3]}°{digits[4:]}'"
                        elif len(digits) == 5:
                            lng = f"东经 {digits[:3]}°{digits[3:]}'"
                        break
            if lng:
                break
                
    if lat and lng:
        return f"{lat}, {lng}"
    elif lat:
        return lat
    elif lng:
        return lng
    return ""

def get_tolls_from_pdf(pdf_path, legs):
    if not os.path.exists(pdf_path):
        print(f"Toll PDF not found at {pdf_path}. All tolls set to 0.0.")
        return [0.0] * len(legs)
        
    pdf_text = ""
    try:
        from pypdf import PdfReader
        reader = PdfReader(pdf_path)
        for page in reader.pages:
            pdf_text += page.extract_text() or ""
    except Exception as e:
        print(f"Failed to read PDF using pypdf: {e}")
        try:
            import pdfplumber
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    pdf_text += page.extract_text() or ""
        except Exception as e2:
            print(f"Failed to read PDF using pdfplumber: {e2}")
            
    if not pdf_text:
        print("Could not extract any text from toll PDF. All tolls set to 0.0.")
        return [0.0] * len(legs)
        
    lines = [line.strip() for line in pdf_text.split('\n') if line.strip()]
    segments = []
    i = 0
    while i < len(lines):
        line = lines[i]
        match = re.match(r'^(\d+)\s+(\d{8})', line)
        if match:
            seg_idx = int(match.group(1))
            date_str = match.group(2)
            i += 1
            seg_lines = []
            while i < len(lines) and not re.match(r'^(\d+)\s+(\d{8})', lines[i]) and "共16段行程" not in lines[i]:
                seg_lines.append(lines[i])
                i += 1
            
            numbers = []
            station_text = ""
            for sl in seg_lines:
                if "至" in sl:
                    station_text += " 至 "
                elif re.search(r'\d+\.\d+', sl):
                    numbers.extend(re.findall(r'\d+\.\d+', sl))
                else:
                    station_text += sl.replace(" ", "")
            
            amount = 0.0
            if numbers:
                amount = float(numbers[0])
            
            segments.append({
                "index": seg_idx,
                "date": f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}",
                "stations": station_text,
                "amount": amount
            })
        else:
            i += 1
            
    # Pair adjacent segments
    pdf_tolls = []
    idx = 0
    while idx < len(segments):
        seg = segments[idx]
        if seg["amount"] == 9.50:
            if idx + 1 < len(segments) and segments[idx + 1]["amount"] != 9.50 and segments[idx + 1]["date"] == seg["date"]:
                pdf_tolls.append(seg["amount"] + segments[idx + 1]["amount"])
                idx += 2
            else:
                pdf_tolls.append(seg["amount"])
                idx += 1
        else:
            if idx + 1 < len(segments) and segments[idx + 1]["amount"] == 9.50 and segments[idx + 1]["date"] == seg["date"]:
                pdf_tolls.append(seg["amount"] + segments[idx + 1]["amount"])
                idx += 2
            else:
                pdf_tolls.append(seg["amount"])
                idx += 1
                
    # Match the paired PDF tolls with Hefei legs chronologically
    hefei_leg_indices = []
    for idx_leg, leg in enumerate(legs):
        sc = leg['start_city'] or ""
        ec = leg['end_city'] or ""
        if ("合肥" in sc and "南京" in ec) or ("南京" in sc and "合肥" in ec):
            hefei_leg_indices.append(idx_leg)
            
    tolls = [0.0] * len(legs)
    if len(hefei_leg_indices) == len(pdf_tolls):
        for i_hefei, orig_idx in enumerate(hefei_leg_indices):
            tolls[orig_idx] = pdf_tolls[i_hefei]
    else:
        print(f"Warning: Hefei legs count ({len(hefei_leg_indices)}) does not match PDF tolls count ({len(pdf_tolls)}).")
        limit = min(len(hefei_leg_indices), len(pdf_tolls))
        for i_hefei in range(limit):
            tolls[hefei_leg_indices[i_hefei]] = pdf_tolls[i_hefei]
            
    return tolls

def run_auto_reimbursement(excel_file='用车费用明细.xlsx', img_dir='photos'):
    # Try to load config.json
    config = {
        "excel_file": excel_file,
        "template_file": "用车费用明细_template.xlsx",
        "img_dir": img_dir,
        "title": "2026年4月份出差车辆行车记录表",
        "license_plate": "苏A859EB",
        "department_reimburser": "报销部门：亥客技术                     报销人：郭良志",
        "document_date": "制表日期："
    }
    
    config_path = 'config.json'
    if os.path.exists(config_path):
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                user_config = json.load(f)
                config.update(user_config)
            print(f"Loaded configuration from {config_path}")
        except Exception as e:
            print(f"Warning: Failed to load {config_path}: {e}")
            
    excel_file = config.get("excel_file", excel_file)
    img_dir = config.get("img_dir", img_dir)
    template_file = config.get("template_file", "用车费用明细_template.xlsx")
    
    # Dynamically resolve document_date to use current system date
    today = datetime.date.today()
    current_date_str = f"{today.year}年{today.month}月{today.day}日"
    configured_date = config.get("document_date", "")
    
    if configured_date:
        pattern_cn = r'\d{4}年\d{1,2}月\d{1,2}日'
        pattern_dash = r'\d{4}-\d{2}-\d{2}'
        pattern_dot = r'\d{4}\.\d{2}\.\d{2}'
        
        if re.search(pattern_cn, configured_date):
            document_date_final = re.sub(pattern_cn, current_date_str, configured_date)
        elif re.search(pattern_dash, configured_date):
            dash_str = today.strftime('%Y-%m-%d')
            document_date_final = re.sub(pattern_dash, dash_str, configured_date)
        elif re.search(pattern_dot, configured_date):
            dot_str = today.strftime('%Y.%m.%d')
            document_date_final = re.sub(pattern_dot, dot_str, configured_date)
        else:
            if configured_date.endswith("：") or configured_date.endswith(":"):
                document_date_final = f"{configured_date}{current_date_str}"
            elif "制表" in configured_date and not any(c in configured_date for c in [":", "："]):
                document_date_final = f"{configured_date}：{current_date_str}"
            else:
                document_date_final = f"{configured_date}{current_date_str}"
    else:
        document_date_final = f"制表日期：{current_date_str}"
        
    config["document_date"] = document_date_final
    
    # Copy fresh template to excel_file if template exists
    if os.path.exists(template_file):
        try:
            import shutil
            shutil.copy(template_file, excel_file)
            print(f"Copied fresh template from {template_file} to {excel_file}")
        except Exception as e:
            print(f"Warning: Failed to copy template {template_file} to {excel_file}: {e}")

    # Store config in an attribute or keep local variables
    run_auto_reimbursement.config = config

    if not os.path.exists(img_dir) and img_dir == 'photos':
        img_dir = '.'
    # 1. Gather all JPG images
    image_files = sorted([os.path.join(img_dir, f) for f in os.listdir(img_dir) if f.lower().endswith('.jpg')])
    if not image_files:
        print("No JPG images found in the directory.")
        return False
        
    print(f"Found {len(image_files)} images. Initializing EasyOCR...")
    import easyocr
    reader = easyocr.Reader(['ch_sim', 'en'])
    
    # 2. Extract OCR and metadata
    images_data = []
    for f in image_files:
        print(f"Extracting: {f}")
        loc = get_location_from_exif(f)
        result = reader.readtext(f)
        ocr_texts = [res[1] for res in result]
        
        date_str = None
        city = None
        
        for t in ocr_texts:
            t_clean = t.strip()
            date_match = re.search(r'(\d{4})[.-](\d{2})[.-](\d{2})', t_clean)
            if date_match:
                date_str = f"{date_match.group(1)}.{date_match.group(2)}.{date_match.group(3)}"
            if not loc and '市' in t_clean:
                loc = t_clean
                
        if loc:
            parts = re.split(r'[·:：\-]', loc)
            city = parts[0].strip()
            
        odo = get_clean_odometer(ocr_texts)
        gps_raw = parse_gps(ocr_texts)
        
        date_obj = None
        if date_str:
            parts = date_str.split('.')
            date_obj = datetime.date(int(parts[0]), int(parts[1]), int(parts[2]))
            
        images_data.append({
            "file": f,
            "date": date_str,
            "date_obj": date_obj,
            "city": city,
            "location": loc,
            "odo": odo,
            "gps_raw": gps_raw,
            "ocr_raw": ocr_texts
        })

    # Reconcile missing GPS components by matching locations
    gps_by_location = {}
    for item in images_data:
        loc = item['location']
        if loc and item['gps_raw'] and ',' in item['gps_raw']:
            gps_by_location[loc] = item['gps_raw']
            
    for item in images_data:
        loc = item['location']
        if loc and loc in gps_by_location:
            item['gps'] = gps_by_location[loc]
        else:
            item['gps'] = item['gps_raw']

    # Sort images by odometer readings
    images_data.sort(key=lambda x: x['odo'] if x['odo'] else 0)
    
    # 3. Construct legs
    legs = []
    for i in range(len(images_data) - 1):
        img_start = images_data[i]
        img_end = images_data[i+1]
        
        if img_start['city'] != img_end['city']:
            legs.append({
                "date": img_end['date_obj'],
                "start_city": img_start['city'],
                "end_city": img_end['city'],
                "start_odo": img_start['odo'],
                "end_odo": img_end['odo']
            })
            
    print(f"Generated {len(legs)} legs.")
    if not legs:
        print("No travel legs detected from location transitions.")
        return False
        
    # 4. Open and update Excel
    if not os.path.exists(excel_file):
        print(f"Excel file {excel_file} not found.")
        return False
        
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    
    # Write custom user configuration to header cells
    config = run_auto_reimbursement.config
    sheet.cell(row=1, column=1, value=config.get("title", "2026年4月份出差车辆行车记录表"))
    sheet.cell(row=2, column=2, value=config.get("license_plate", "苏A859EB"))
    sheet.cell(row=2, column=3, value=config.get("department_reimburser", "报销部门：亥客技术                     报销人：郭良志"))
    sheet.cell(row=2, column=9, value=config.get("document_date", "制表日期：2026年5月28日"))
    
    # Locate '合计：'
    total_row_idx = None
    for r in range(5, 100):
        if sheet.cell(row=r, column=2).value == "合计：":
            total_row_idx = r
            break
            
    if not total_row_idx:
        print("Could not find '合计：' row in column B.")
        return False
        
    num_legs = len(legs)
    existing_data_rows = total_row_idx - 5
    
    # Clean up merged cells in the data and shifted region (Row 13 and below)
    ranges_to_remove = []
    for r_range in list(sheet.merged_cells.ranges):
        if r_range.min_row >= 13:
            ranges_to_remove.append(r_range)
            
    for r_range in ranges_to_remove:
        sheet.merged_cells.remove(r_range)
        
    # Save and reload workbook to reset the read-only MergedCell objects in openpyxl cache
    wb.save(excel_file)
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    
    if num_legs > existing_data_rows:
        rows_to_insert = num_legs - existing_data_rows
        print(f"Inserting {rows_to_insert} rows at Row {total_row_idx}...")
        
        # Save styling from template row 5
        template_row = 5
        row_styles = {}
        for col_idx in range(1, 13):
            cell = sheet.cell(row=template_row, column=col_idx)
            row_styles[col_idx] = {
                "font": cell.font,
                "fill": cell.fill,
                "border": cell.border,
                "alignment": cell.alignment,
                "number_format": cell.number_format
            }
            
        sheet.insert_rows(total_row_idx, rows_to_insert)
        new_total_row_idx = total_row_idx + rows_to_insert
        
        # Style newly inserted rows
        for r in range(total_row_idx, new_total_row_idx):
            for col_idx in range(1, 13):
                cell = sheet.cell(row=r, column=col_idx)
                style = row_styles[col_idx]
                if style["font"]: cell.font = Font(name=style["font"].name, size=style["font"].size, bold=style["font"].bold, italic=style["font"].italic, color=style["font"].color)
                if style["fill"]: cell.fill = PatternFill(fill_type=style["fill"].fill_type, start_color=style["fill"].start_color, end_color=style["fill"].end_color)
                if style["border"]: cell.border = Border(left=style["border"].left, right=style["border"].right, top=style["border"].top, bottom=style["border"].bottom)
                if style["alignment"]: cell.alignment = Alignment(horizontal=style["alignment"].horizontal, vertical=style["alignment"].vertical, wrap_text=style["alignment"].wrap_text)
                cell.number_format = style["number_format"]
    else:
        new_total_row_idx = total_row_idx
        
    # Write leg data
    pdf_path = os.path.join(os.path.dirname(os.path.abspath(excel_file)), 'trans', 'trans.pdf')
    tolls = get_tolls_from_pdf(pdf_path, legs)
    
    for idx, leg in enumerate(legs):
        r = 5 + idx
        sheet.cell(row=r, column=1, value=leg['date'])
        sheet.cell(row=r, column=3, value=leg['start_city'])
        sheet.cell(row=r, column=4, value=leg['start_odo'])
        sheet.cell(row=r, column=5, value=leg['end_city'])
        sheet.cell(row=r, column=6, value=leg['end_odo'])
        sheet.cell(row=r, column=8, value=f"=F{r}-D{r}")
        sheet.cell(row=r, column=9, value=f"=H{r}*0.8")
        sheet.cell(row=r, column=11, value=tolls[idx] if idx < len(tolls) else 0.0)
        
    # Re-merge cells shifted down
    sig_row = new_total_row_idx + 1
    sheet.merge_cells(f"C{sig_row}:D{sig_row}") # 领导签名
    
    # Total row formulas
    sheet.cell(row=new_total_row_idx, column=8, value=f"=SUM(H5:H{new_total_row_idx-1})")
    sheet.cell(row=new_total_row_idx, column=9, value=f"=SUM(I5:I{new_total_row_idx-1})")
    sheet.cell(row=new_total_row_idx, column=11, value=f"=SUM(K5:K{new_total_row_idx-1})")
    
    # Locate summary section
    summary_header_row = None
    for r in range(new_total_row_idx, new_total_row_idx + 10):
        if sheet.cell(row=r, column=2).value == "总里程数":
            summary_header_row = r
            break
            
    if summary_header_row:
        summary_data_row = summary_header_row + 1
        sheet.cell(row=summary_data_row, column=2, value=f"=H{new_total_row_idx}") # 总里程数
        sheet.cell(row=summary_data_row, column=3, value=f"=H{new_total_row_idx}") # 实报里程数
        sheet.cell(row=summary_data_row, column=5, value=f"=K{new_total_row_idx}") # 过路费
        sheet.cell(row=summary_data_row, column=7, value=f"=C{summary_data_row}*0.8+E{summary_data_row}+F{summary_data_row}") # 报销金额
        sheet.merge_cells(f"G{summary_data_row}:I{summary_data_row}") # 报销金额合并
        photo_log_start = summary_data_row + 2
    else:
        photo_log_start = new_total_row_idx + 6
        
    # 5. Compress and Insert Image Log starting from photo_log_start
    print(f"Inserting Photo Log starting from Row {photo_log_start}...")
    
    font_bold = Font(name="宋体", size=11, bold=True)
    font_regular = Font(name="宋体", size=11)
    fill_header = PatternFill(fill_type="solid", start_color="F2F2F2", end_color="F2F2F2")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    thin_side = Side(border_style="thin", color="D9D9D9")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # Header
    sheet.row_dimensions[photo_log_start].height = 25
    headers = {
        2: "车辆仪表照片",
        3: "对应图片文件名",
        4: "仪表盘总里程 (km)",
        5: "照片水印地点",
        6: "经纬度",
        7: "拍摄日期"
    }
    for col, text in headers.items():
        cell = sheet.cell(row=photo_log_start, column=col, value=text)
        cell.font = font_bold
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all
        
    sheet.column_dimensions['B'].width = 22
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 30
    sheet.column_dimensions['F'].width = 30
    sheet.column_dimensions['G'].width = 15
    
    # Process each image
    comp_dir = os.path.join(os.path.dirname(os.path.abspath(excel_file)), 'compressed')
    os.makedirs(comp_dir, exist_ok=True)
    
    for idx, item in enumerate(images_data):
        r = photo_log_start + 1 + idx
        sheet.row_dimensions[r].height = 165
        
        img_name = os.path.basename(item['file'])
        comp_path = os.path.join(comp_dir, img_name)
        
        # Compress to 15% width/height
        with Image.open(item['file']) as pil_img:
            width, height = pil_img.size
            new_width = int(width * 0.15)
            new_height = int(height * 0.15)
            resized_img = pil_img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            resized_img.save(comp_path, quality=70)
            
        excel_img = OpenpyxlImage(comp_path)
        sheet.add_image(excel_img, f"B{r}")
        
        # Metadata
        c_file = sheet.cell(row=r, column=3, value=img_name)
        c_odo = sheet.cell(row=r, column=4, value=item['odo'])
        c_loc = sheet.cell(row=r, column=5, value=item['location'])
        c_gps = sheet.cell(row=r, column=6, value=item['gps'])
        c_date = sheet.cell(row=r, column=7, value=item['date_obj'])
        
        for col in [2, 3, 4, 5, 6, 7]:
            cell = sheet.cell(row=r, column=col)
            cell.font = font_regular
            cell.border = border_all
            if col in [5, 6]:
                cell.alignment = align_left
            else:
                cell.alignment = align_center
                
        c_date.number_format = 'yyyy-mm-dd'
        c_odo.number_format = '#,##0'
        
    wb.save(excel_file)
    print(f"Successfully processed reimbursement flow and updated {excel_file}!")
    return True

if __name__ == '__main__':
    run_auto_reimbursement()
